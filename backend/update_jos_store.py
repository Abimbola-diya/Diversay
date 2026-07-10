import logging
import openpyxl
from database import SessionLocal
from models import Product, Store, StoreInventory, UnitType, ProductCategory

logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

file_path = "/home/abimbola/Desktop/Diversay_bootstrapped/DSLPC INVENTORY REPORT FOR  PRODUCTION PLAN WEEK 2 (4TH-10TH JAN 2026)._ (1).xlsx"

# Mapping from Excel product name to database-friendly product name and unit
MAPPING = {
    # DSL Products
    "Bicofim -500grm": ("Bicofim 500gm", UnitType.PCS),
    "Biokleen 1LITRE": ("Biokleen 1L", UnitType.PCS),
    "Biokleen 20 Litres.": ("Biokleen 20L", UnitType.KEG),
    "Biokleen 500ml": ("Biokleen 500ml", UnitType.PCS),
    "Citramax 1ltr": ("Citramax 1L", UnitType.PCS),
    "Citramax 20ltrs": ("Citramax 20L", UnitType.KEG),
    "Neodine 1 Ltr ": ("Neodine 1L", UnitType.PCS),
    "Neodine 20 Litres": ("Neodine 20L", UnitType.KEG),
    "Neodine 500ml": ("Neodine 500ml", UnitType.PCS),
    "Terminator 1ltr ": ("Terminator 1L", UnitType.PCS),
    "Terminator 20 ltr ": ("Terminator 20L", UnitType.KEG),
    "Tetrasan 1ltr": ("Tetrasan 1L", UnitType.PCS),
    "Tetrasan 20 Litres": ("Tetrasan 20L", UnitType.KEG),
    "Tetrasan 500ml": ("Tetrasan 500ml", UnitType.PCS),
    "Viracid-S - 1kg": ("Viracid-s 1kg", UnitType.PCS),
    "Viracid-S 5.0kg": ("Viracid-s 5kg", UnitType.PCS),
    "Viracid-S 20 kg": ("Viracid-s 20kg", UnitType.DRUM),
    "Wyldox 500GM": ("Wyldox 500gm", UnitType.PCS),
    "Wyldox 100GM": ("Wyldox 100gm", UnitType.PCS),
    "Wyldox 20GM": ("Wyldox 20gm", UnitType.PCS),
    "WYLDOX TABLET (10G *100PCS)": ("Wyldox tablet", UnitType.PCS),
    "Biokleen Concentrate 25ltr": ("Biokleen Concentrate 25L", UnitType.PCS),
    "Citramax Concentrate - 25ltr": ("Citramax Concentrate 25L", UnitType.PCS),
    "Neodine Concentrate": ("Neodine Concentrate", UnitType.PCS),
    "Terminator 111 Concentrate 25ltr": ("Terminator III Concentrate 25L", UnitType.PCS),
    "Tetrasan Concentrate 25ltr": ("Tetrasan Concentrate 25L", UnitType.PCS),
    "Avertin Oralsolution 1% 100ml": ("Avertin 100ml", UnitType.PCS),
    "Ds-LivorTon Liquid-500ml": ("Ds livorton 500ml", UnitType.PCS),
    "DS-Siproxin 10%-100gm": ("DS-Siproxin 10% 100gm", UnitType.PCS),
    "Megadox-N 100gm": ("Megadox-N 100gm", UnitType.PCS),
    "Oxy-100-FS - 10kg": ("Oxy-100-FS 10kg", UnitType.PCS),
    "QUINROCIN ORAL 1 LTRE": ("Quinrocin 1L", UnitType.PCS),
    "TYLOTON-100G": ("Tyloton 100gm", UnitType.PCS),
    "Ultracal - D - 500ml": ("Ultracal-d 500ml", UnitType.PCS),
    "Ultra -TM Plus 500ml": ("Ultra-TM Plus 500ml", UnitType.PCS),
    "Ultravite-M/WS-100gm ": ("Ultravite-M/WS 100gm", UnitType.PCS),
    "Acifed-Fs- 20kg": ("Acifed-FS (20kg)", UnitType.BAG),
    "A-GEEMIX FORTE": ("A gee mix forte", UnitType.BAG),
    "Curatox-Fs": ("Curatox (25kg)", UnitType.BAG),
    "HERBAN 20kg": ("Herban (20kg)", UnitType.BAG),
    "KIKSTART-CXL 1KG": ("Kikstart-CXL 1kg", UnitType.PCS),
    "Kolin Powder - 25kg": ("Kolin Powder (25kg)", UnitType.BAG),
    "MAXIZYME": ("Maxizyme", UnitType.DRUM),
    "Natchol - 25kg": ("Natchol (25kg)", UnitType.BAG),
    "Phytocee - 25kg": ("Phytocee powder (25kg)", UnitType.BAG),
    "Phytocee Cool - 1kg .": ("Phytocee Cool 1kg", UnitType.PCS),
    "Phytonin - 25kg .": ("Phytonin (25kg)", UnitType.BAG),
    "Stodi - 25kg": ("Stodi (25kg)", UnitType.BAG),
    "Topicure Spray 250ml (25/ctn)": ("Topicure 250ml", UnitType.PCS),
    "ULTRASIL-TCF 25KG": ("Ultrasil-TCF (25kg)", UnitType.BAG),
    "Zigbir Powder -25kg": ("Zigbir Powder (25kg)", UnitType.BAG),
    "Zigbir Liquid 1 ltre.": ("Zigbir 1L", UnitType.PCS),

    # DSLP Products
    "AMPROLIUM 30 - 100GRM": ("Amprolium30 100gm", UnitType.PCS),
    "COXSTOP  100GRM": ("Coxstop 100gm", UnitType.PCS),
    "COXSTOP  30GRM": ("Coxstop 30gm", UnitType.PCS),
    "DISULTRIM  100GRM": ("Dlsultrim 100gm", UnitType.PCS),
    "DISULTRIM 50GRM": ("Dlsultrim 50gm", UnitType.PCS),
    "DIVERCAL-D": ("Divercal-d 1L", UnitType.PCS),
    "DIVERCIPRO 1 LTR": ("Divercipro 1L", UnitType.PCS),
    "DIVERCOOL C 1 Ltr": ("Divercool-c 1L", UnitType.PCS),
    "DIVERCOOL C 25Litres": ("Divercool-c 25L", UnitType.KEG),
    "DIVERCOX 1 LITRE": ("Divercox 1L", UnitType.PCS),
    "DIVERDAZOLE 1 LTR": ("Diverdazole 1L", UnitType.PCS),
    "DIVERFLOX    100GRM": ("Diverflox 100gm", UnitType.PCS),
    "DIVERFUME 1KG": ("Diverfume 1kg", UnitType.PCS),
    "DIVERGEN D  100GRM": ("Divergen-d 100gm", UnitType.PCS),
    "DIVERTAMIN": ("Divertamin 100gm", UnitType.PCS),
    "DIVERVITE 1 LTRE": ("Divervite 1L", UnitType.PCS),
    "DOXINEO 100GRM": ("Doxineo 100gm", UnitType.PCS),
    "DOXINEO 50GRM": ("Doxineo 50gm", UnitType.PCS),
    "ENRO  1 LTR": ("Enrol 1L", UnitType.PCS),
    "GENTYLO - 100 GRM": ("Gentylo 100gm", UnitType.PCS),
    "IVERTMECTIN 1 LTR": ("Divermectin 1L", UnitType.PCS),
    "DIVERSEL E 1 LITRE": ("Diversel-e 1L", UnitType.PCS),
    "DIVERSEL E 25 LITRES": ("Diversel-e 25L", UnitType.KEG),
    "LEVASTAR 10OGRM": ("Levastar 100gm", UnitType.PCS),
    "LEVASTAR 5OGRM": ("Levastar 50gm", UnitType.PCS),
    "TRUSAN LIQUID SOAP 25LTR": ("Trusan 25L", UnitType.KEG),
    "VETODINE 1 LTR": ("Vetodine 1L", UnitType.PCS),
    "VETODINE 25 LTR": ("Vetodine 25L", UnitType.KEG),
}

def parse_excel_data():
    wb = openpyxl.load_workbook(file_path, data_only=True)
    
    extracted_data = []
    
    # 1. Parse DSL WEEK 2
    # JOS is at column index 8 (0-based)
    dsl_sheet = wb["DSL WEEK 2"]
    for idx, row in enumerate(dsl_sheet.iter_rows(values_only=True)):
        if idx < 4:
            continue
        name = row[1]
        unit_str = row[2]
        jos_val = row[8]
        
        if name and not name.startswith("SUB TOTAL") and not name.startswith("SUB TOTA") and name not in ["Biosecurity", "Concentrates", "Drugs", "Equipment", "Packaging", "Premixes"]:
            # Check if this is a mapped product
            db_name, default_unit = MAPPING.get(name, (name, UnitType.PCS))
            # Determine unit type
            if unit_str:
                unit_str_upper = str(unit_str).upper()
                if "PCS" in unit_str_upper:
                    default_unit = UnitType.PCS
                elif "KEG" in unit_str_upper:
                    default_unit = UnitType.KEG
                elif "BAG" in unit_str_upper:
                    default_unit = UnitType.BAG
                elif "DRUM" in unit_str_upper:
                    default_unit = UnitType.DRUM
            
            # Map JOS value
            stock_val = 0.0
            if jos_val is not None:
                try:
                    stock_val = float(jos_val)
                except ValueError:
                    pass
            
            extracted_data.append({
                "excel_name": name,
                "db_name": db_name,
                "unit": default_unit,
                "stock": stock_val,
                "type": "DSL",
                "original_val": jos_val
            })
            
    # 2. Parse DSLP FG WK 2
    # JOS is at column index 9 (0-based)
    dslp_sheet = wb["DSLP FG WK 2"]
    for idx, row in enumerate(dslp_sheet.iter_rows(values_only=True)):
        if idx < 3:
            continue
        name = row[2] # product name is at index 2
        jos_val = row[9]
        
        if name and not name.startswith("SUB TOTAL") and not name.startswith("TOTAL") and name != "FINISHED PRODUCTS":
            db_name, default_unit = MAPPING.get(name, (name, UnitType.PCS))
            stock_val = 0.0
            if jos_val is not None:
                try:
                    stock_val = float(jos_val)
                except ValueError:
                    pass
            
            extracted_data.append({
                "excel_name": name,
                "db_name": db_name,
                "unit": default_unit,
                "stock": stock_val,
                "type": "DSLP",
                "original_val": jos_val
            })
            
    return extracted_data

def update_jos_store():
    db = SessionLocal()
    try:
        # Find Jos Store
        jos_store = db.query(Store).filter(
            Store.name.ilike("%Jos Store%"),
            Store.is_deleted == False
        ).first()
        
        if not jos_store:
            logger.error("Jos Store not found in database!")
            return
            
        logger.info(f"Found Jos Store with ID: {jos_store.id}")
        
        data = parse_excel_data()
        
        print("\n--- UPDATING JOS STORE INVENTORY ---")
        for item in data:
            # 1. Get or create product
            product = db.query(Product).filter(
                Product.name.ilike(item["db_name"]),
                Product.is_deleted == False
            ).first()
            
            if not product:
                logger.info(f"Creating product: {item['db_name']} with default unit: {item['unit']}")
                product = Product(
                    name=item["db_name"],
                    category=ProductCategory.OTHER,
                    default_unit=item["unit"],
                    unit_price=0.0
                )
                db.add(product)
                db.commit()
                db.refresh(product)
                
            # 2. Update stock for Jos Store
            inv = db.query(StoreInventory).filter(
                StoreInventory.store_id == jos_store.id,
                StoreInventory.product_id == product.id
            ).first()
            
            if not inv:
                logger.info(f"Creating inventory record for {product.name} at Jos Store with stock {item['stock']}")
                inv = StoreInventory(
                    store_id=jos_store.id,
                    product_id=product.id,
                    stock=item["stock"]
                )
                db.add(inv)
            else:
                logger.info(f"Updating stock for {product.name} at Jos Store: {inv.stock} -> {item['stock']}")
                inv.stock = item["stock"]
                
        db.commit()
        logger.info("Jos Store inventory updated successfully!")
        
        # Print a report of the products updated, indicating type (DSL/DSLP) and stock value
        print("\n=== INVENTORY IMPORT SUMMARY FOR JOS STORE ===")
        print(f"{'Type':<6} | {'Product Name':<35} | {'Unit':<10} | {'Stock':<10} | {'Original Cell Value'}")
        print("-" * 80)
        for item in sorted(data, key=lambda x: (x["type"], x["db_name"])):
            print(f"{item['type']:<6} | {item['db_name']:<35} | {item['unit'].value:<10} | {item['stock']:<10} | {item['original_val']}")
            
    except Exception as e:
        logger.error(f"Error updating Jos Store inventory: {e}")
        db.rollback()
    finally:
        db.close()

if __name__ == "__main__":
    update_jos_store()
