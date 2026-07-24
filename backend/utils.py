from datetime import datetime, timezone, timedelta
from models import Order, OrderStatus
from sqlalchemy.orm import Session
from difflib import SequenceMatcher
from openpyxl import load_workbook
from typing import List, Tuple
import io
import json

def to_naive(dt: datetime) -> datetime:
    """Normalize a datetime to UTC naive datetime."""
    if dt is None:
        return None
    if dt.tzinfo is not None:
        return dt.astimezone(timezone.utc).replace(tzinfo=None)
    return dt

def calculate_order_status(order: Order) -> OrderStatus:
    """Calculate order status based on timestamps and delivery status.
    Defaults expected_delivery_time to dispatch_time + 48 hours if omitted.
    """
    actual = to_naive(order.actual_delivery_time)
    expected = to_naive(order.expected_delivery_time)
    dispatch = to_naive(order.dispatch_time)
    created = to_naive(order.created_at) if hasattr(order, 'created_at') else None

    if expected is None:
        if dispatch:
            expected = dispatch + timedelta(hours=48)
        elif created:
            expected = created + timedelta(hours=48)
    
    if actual:
        if expected is None:
            return OrderStatus.DELIVERED_ON_TIME
        if actual <= expected:
            return OrderStatus.DELIVERED_ON_TIME
        else:
            return OrderStatus.DELIVERED_LATE
    else:
        if not dispatch:
            return OrderStatus.DRAFT
        
        now = datetime.utcnow()
        if expected is None:
            return OrderStatus.IN_TRANSIT
        if now <= expected:
            return OrderStatus.IN_TRANSIT
        else:
            return OrderStatus.DELAYED

def calculate_delivery_duration(order: Order) -> int:
    """Calculate delivery duration in hours between dispatch and actual delivery."""
    dispatch = to_naive(order.dispatch_time)
    actual = to_naive(order.actual_delivery_time)
    if dispatch and actual:
        delta = actual - dispatch
        return int(delta.total_seconds() / 3600)
    return None

def generate_order_number(db: Session) -> str:
    """Generate unique order number in format: DSL-YYYY-XXXX"""
    year = datetime.utcnow().year
    
    last_order = db.query(Order).filter(
        Order.order_number.ilike(f"DSL-{year}-%")
    ).order_by(Order.id.desc()).first()
    
    if last_order:
        seq = int(last_order.order_number.split("-")[-1])
        seq += 1
    else:
        seq = 1
    
    order_number = f"DSL-{year}-{seq:04d}"
    return order_number

def _trigram_similarity(a: str, b: str) -> float:
    """Calculate trigram (3-gram) overlap between two strings."""
    if len(a) < 3 or len(b) < 3:
        # Fall back to bigrams for short strings
        n = 2
    else:
        n = 3
    
    def ngrams(s, n):
        return set(s[i:i+n] for i in range(len(s) - n + 1))
    
    grams_a = ngrams(a, n)
    grams_b = ngrams(b, n)
    
    if not grams_a or not grams_b:
        return 0.0
    
    intersection = grams_a & grams_b
    union = grams_a | grams_b
    return len(intersection) / len(union)  # Jaccard similarity


def fuzzy_search_customers(query: str, customers: List[dict], threshold: float = 0.15) -> List[dict]:
    """
    Smart search customers by name using multi-signal scoring:
      1. Exact prefix match (name starts with query)        → +5.0
      2. Any word in the name starts with the query          → +3.0
      3. Substring containment (query found inside name)     → +2.0
      4. Trigram similarity for typo tolerance                → 0.0–1.0
    Results are sorted by total score descending.
    """
    if not query:
        return []
    
    query_lower = query.lower().strip()
    results = []
    
    for customer in customers:
        name_lower = customer['name'].lower()
        score = 0.0
        
        # Signal 1: Name starts with the query (strongest signal)
        if name_lower.startswith(query_lower):
            score += 5.0
        
        # Signal 2: Any word in the name starts with the query
        words = name_lower.split()
        if any(w.startswith(query_lower) for w in words):
            score += 3.0
        
        # Signal 3: Substring containment
        if query_lower in name_lower:
            score += 2.0
        
        # Signal 4: Trigram similarity (catches typos)
        tri_sim = _trigram_similarity(query_lower, name_lower)
        score += tri_sim
        
        if score >= threshold:
            results.append((customer, score))
    
    results.sort(key=lambda x: (-x[1], x[0]['name']))
    return [r[0] for r in results]

def parse_excel_customers(file_content: bytes) -> List[dict]:
    """Parse customers from Excel file."""
    try:
        workbook = load_workbook(io.BytesIO(file_content))
        sheet = workbook.active
        
        headers = {}
        for col_idx, cell in enumerate(sheet[1], 1):
            if cell.value:
                headers[cell.value.lower().strip()] = col_idx
        
        expected_cols = {
            'name': ['name', 'customer name', 'customer_name'],
            'address': ['address', 'delivery address'],
            'city': ['city'],
            'state': ['state'],
            'contact_number': ['contact', 'contact number', 'contact_number', 'phone'],
            'email': ['email', 'email address']
        }
        
        col_mapping = {}
        for field, alternatives in expected_cols.items():
            for header_key, col_idx in headers.items():
                if any(alt in header_key for alt in alternatives):
                    col_mapping[field] = col_idx
                    break
        
        customers = []
        for row_idx, row in enumerate(sheet.iter_rows(min_row=2, values_only=False), 2):
            customer = {}
            
            for field, col_idx in col_mapping.items():
                cell_value = sheet.cell(row=row_idx, column=col_idx).value
                customer[field] = cell_value if cell_value else None
            
            if customer.get('name'):
                customers.append(customer)
        
        return customers
    
    except Exception as e:
        raise ValueError(f"Error parsing Excel file: {str(e)}")

def get_order_with_details(order: Order) -> dict:
    """Convert order object to dict with calculated fields and related data."""
    status = calculate_order_status(order)
    duration = calculate_delivery_duration(order)
    
    # Calculate logistics costs
    fuel_cost = getattr(order, "fuel_cost", 0.0) or 0.0
    waybill_cost = getattr(order, "waybill_cost", 0.0) or 0.0
    other_costs = getattr(order, "other_costs", []) or []
    
    total_other = 0.0
    if isinstance(other_costs, list):
        for o in other_costs:
            if isinstance(o, dict):
                try:
                    total_other += float(o.get("amount", 0.0) or 0.0)
                except ValueError:
                    pass
                    
    total_amount = float(fuel_cost) + float(waybill_cost) + total_other
    total_qty = sum(item.quantity for item in order.line_items)
    avg_logistics_rate = total_amount / total_qty if total_qty > 0 else 0.0
    
    customer_dict = None
    if order.customer:
        customer_dict = {
            "id": order.customer.id,
            "name": order.customer.name,
            "address": order.customer.address,
            "city": order.customer.city,
            "state": order.customer.state,
            "contact_number": order.customer.contact_number,
            "email": order.customer.email,
            "created_at": order.customer.created_at,
            "updated_at": order.customer.updated_at
        }

    user_dict = None
    if order.created_by_user:
        user_dict = {
            "id": order.created_by_user.id,
            "email": order.created_by_user.email,
            "full_name": order.created_by_user.full_name,
            "role": order.created_by_user.role.value if order.created_by_user.role else None,
            "is_active": order.created_by_user.is_active,
            "created_at": order.created_by_user.created_at,
            "requesting_admin": order.created_by_user.requesting_admin
        }

    def serialize_line_item(item):
        return {
            "id": item.id,
            "product_id": item.product_id,
            "product_name": item.product.name if item.product else "Unknown",
            "product_brand": item.product.brand if item.product else None,
            "quantity": item.quantity,
            "unit": item.unit.value,
            "unit_price": avg_logistics_rate,
            "total_price": avg_logistics_rate * item.quantity,
            "created_at": item.created_at
        }

    # Build reference cards response
    ref_cards_list = []
    if hasattr(order, 'reference_cards') and order.reference_cards:
        for card in order.reference_cards:
            ref_cards_list.append({
                "id": card.id,
                "invoice_number": card.invoice_number,
                "waybill_number": card.waybill_number,
                "brand": card.brand or "DSL",
                "line_items": [serialize_line_item(item) for item in card.line_items]
            })

    return {
        "id": order.id,
        "order_number": order.order_number,
        "waybill_number": order.waybill_number,
        "invoice_number": order.invoice_number,
        "customer_id": order.customer_id,
        "customer_name": order.customer.name if order.customer else None,
        "customer_state": order.customer.state if order.customer else None,
        "customer_address": order.customer.address if order.customer else None,
        "source_store_id": order.source_store_id,
        "destination_store_id": order.destination_store_id,
        "source_store_name": order.source_store.name if order.source_store else None,
        "destination_store_name": order.destination_store.name if order.destination_store else None,
        "destination_store_city": order.destination_store.city if order.destination_store else None,
        "destination_store_state": order.destination_store.state if order.destination_store else None,
        "destination_store_address": order.destination_store.address if order.destination_store else None,
        "source_store_is_central": order.source_store.is_central if order.source_store else False,
        "destination_store_is_central": order.destination_store.is_central if order.destination_store else False,
        "dispatch_time": order.dispatch_time,
        "expected_delivery_time": order.expected_delivery_time,
        "actual_delivery_time": order.actual_delivery_time,
        "delivery_duration": duration,
        "order_status": status,
        "notes": order.notes,
        "driver_name": order.driver_name,
        "vehicle_number": order.vehicle_number,
        "fuel_cost": fuel_cost,
        "waybill_cost": waybill_cost,
        "other_costs": other_costs,
        "created_by_id": order.created_by_id,
        "created_by_name": order.created_by_user.full_name if order.created_by_user else None,
        "created_by_user": user_dict,
        "customer": customer_dict,
        "created_at": order.created_at,
        "updated_at": order.updated_at,
        "total_amount": total_amount,
        "line_items": [serialize_line_item(item) for item in order.line_items],
        "reference_cards": ref_cards_list
    }

def calculate_hours_overdue(expected_time: datetime) -> int:
    """Calculate how many hours past expected delivery time."""
    expected = to_naive(expected_time)
    if expected is None:
        return 0
    now = datetime.utcnow()
    delta = now - expected
    return int(delta.total_seconds() / 3600)
