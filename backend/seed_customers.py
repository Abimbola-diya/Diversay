"""
One-time script to seed customers from the Excel file into the Supabase database.
"""
import sys
import os
sys.path.append(os.path.dirname(__file__))

import openpyxl
from database import SessionLocal
from models import Customer
import logging

logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

EXCEL_PATH = os.path.join(os.path.dirname(__file__), "DSL_DSLP CUSTOMER_2026.xlsx")

def seed_customers():
    db = SessionLocal()
    try:
        wb = openpyxl.load_workbook(EXCEL_PATH)
        ws = wb[wb.sheetnames[0]]

        added = 0
        skipped = 0

        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True):
            ledger_name, address, state, city, contact_no, email_id = row

            # Skip empty rows
            if not ledger_name:
                continue

            name = str(ledger_name).strip()
            address_str = str(address).strip() if address else None
            state_str = str(state).strip() if state else None
            city_str = str(city).strip() if city else None
            # Clean contact number: remove spaces and non-digit chars
            if contact_no:
                raw = str(contact_no).strip().replace(' ', '').replace('-', '')
                # If it's a float like 8.032e9, convert properly
                try:
                    raw = str(int(float(raw)))
                except (ValueError, OverflowError):
                    pass
                contact_str = raw if raw else None
            else:
                contact_str = None
            email_str = str(email_id).strip() if email_id else None

            # Check if customer already exists (by name, case-insensitive)
            existing = db.query(Customer).filter(
                Customer.name.ilike(name),
                Customer.is_deleted == False
            ).first()

            if existing:
                skipped += 1
                continue

            new_customer = Customer(
                name=name,
                address=address_str,
                state=state_str,
                city=city_str,
                contact_number=contact_str,
                email=email_str
            )
            db.add(new_customer)
            added += 1

        db.commit()
        logger.info(f"Seeding complete: {added} customers added, {skipped} duplicates skipped.")
    except Exception as e:
        logger.error(f"Error seeding customers: {e}")
        db.rollback()
        raise
    finally:
        db.close()

if __name__ == "__main__":
    seed_customers()
