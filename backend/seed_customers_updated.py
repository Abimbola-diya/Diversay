"""
Script to update the customer database using the new Excel sheet.
Preserves matching customers in place to maintain integrity of existing orders.
Deletes former customers not in the new sheet, and adds new ones.
"""
import sys
import os
sys.path.append(os.path.dirname(__file__))

import openpyxl
from database import SessionLocal
from models import Customer, Order
import logging

logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

EXCEL_PATH = "/home/abimbola/Desktop/Diversay_bootstrapped/Updated DSL_DSLP Customer 2026 (003).xlsx"

def sync_customers():
    db = SessionLocal()
    try:
        # 1. Load customers from Excel
        wb = openpyxl.load_workbook(EXCEL_PATH)
        ws = wb.active

        excel_customers = {}
        row_count = 0
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True):
            if len(row) < 5:
                continue
            ledger_name, address, state, city, contact_no = row[:5]

            if not ledger_name:
                continue

            name = str(ledger_name).strip()
            name_upper = name.upper()
            
            address_str = str(address).strip() if address else None
            state_str = str(state).strip() if state else None
            city_str = str(city).strip() if city else None

            # Clean contact number
            if contact_no is not None:
                raw = str(contact_no).strip().replace(' ', '').replace('-', '')
                try:
                    # Parse float notation (like 8032646134.0)
                    raw = str(int(float(raw)))
                except (ValueError, OverflowError):
                    pass
                contact_str = raw if raw else None
            else:
                contact_str = None

            # Deduplicate entries in the Excel itself (keep the first or merge, we just use the first/last)
            excel_customers[name_upper] = {
                "name": name,
                "address": address_str,
                "state": state_str,
                "city": city_str,
                "contact_number": contact_str
            }
            row_count += 1

        logger.info(f"Loaded {len(excel_customers)} unique customers from Excel sheet (total rows: {row_count}).")

        # 2. Get referenced customer IDs from orders to prevent deleting them
        referenced_ids = {o.customer_id for o in db.query(Order).all()}
        
        # 3. Load all existing customers from the database
        db_customers = db.query(Customer).all()
        
        added = 0
        updated = 0
        deleted = 0
        skipped_delete = 0

        # Trace DB customers
        for db_cust in db_customers:
            name_upper = db_cust.name.upper()
            if name_upper in excel_customers:
                # Update in place
                data = excel_customers[name_upper]
                db_cust.name = data["name"]
                db_cust.address = data["address"]
                db_cust.state = data["state"]
                db_cust.city = data["city"]
                db_cust.contact_number = data["contact_number"]
                db_cust.email = None # New sheet has no email
                db_cust.is_deleted = False # Ensure not marked as deleted
                updated += 1
                # Remove from excel dict so it's not added as new
                del excel_customers[name_upper]
            else:
                # Not in Excel. Delete it.
                if db_cust.id in referenced_ids:
                    # Safety check: keep it but mark as deleted or warn
                    logger.warning(f"Customer '{db_cust.name}' (ID: {db_cust.id}) is referenced by orders but not in Excel! Marking as is_deleted=True.")
                    db_cust.is_deleted = True
                    skipped_delete += 1
                else:
                    db.delete(db_cust)
                    deleted += 1

        # 4. Insert new customers
        for data in excel_customers.values():
            new_cust = Customer(
                name=data["name"],
                address=data["address"],
                state=data["state"],
                city=data["city"],
                contact_number=data["contact_number"],
                email=None
            )
            db.add(new_cust)
            added += 1

        db.commit()
        logger.info(f"Sync complete: {added} added, {updated} updated, {deleted} deleted, {skipped_delete} skipped delete (marked as deleted).")
        
    except Exception as e:
        logger.error(f"Error during synchronization: {e}")
        db.rollback()
        raise
    finally:
        db.close()

if __name__ == "__main__":
    sync_customers()
